"""
xlsx_generator.py — Generate Excel spreadsheets with live formulas.

Creates XLSX files where calculation cells contain actual Excel formulas
(not static values), making the computation fully transparent and editable.

Key design principle: DATA cells hold raw numbers from exports; FORMULA cells
reference DATA cells using standard Excel expressions. If any data value is
changed, all dependent formulas recompute automatically.
"""
from __future__ import annotations

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

from .export_reader import (
    ComplexityMetricsExport,
    StatisticsExport,
    MessageFlowExport,
)


# ============================================================================
# STYLE CONSTANTS
# ============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="2F5496", size=10)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DATA_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
MONO_FONT = Font(name="Consolas", size=10)
NUMBER_FMT_4 = "0.0000"
NUMBER_FMT_6 = "0.000000"
NUMBER_FMT_PCT = "0.00%"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row: int, max_col: int):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_subheader_row(ws, row: int, max_col: int):
    """Apply subheader styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int, is_formula: bool = False):
    """Apply data or formula styling."""
    cell = ws.cell(row=row, column=col)
    cell.fill = FORMULA_FILL if is_formula else DATA_FILL
    cell.border = THIN_BORDER
    cell.font = MONO_FONT


def _style_result_cell(ws, row: int, col: int):
    """Apply result (green) styling."""
    cell = ws.cell(row=row, column=col)
    cell.fill = RESULT_FILL
    cell.border = THIN_BORDER
    cell.font = Font(bold=True, name="Consolas", size=11)


# ============================================================================
# SHEET BUILDERS
# ============================================================================

def _build_metadata_sheet(
    wb: Workbook,
    experiment_name: str,
    scenario_name: str,
    cm: ComplexityMetricsExport,
):
    """Sheet 1: Experiment and scenario metadata."""
    ws = wb.active
    ws.title = "Metadata"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    data = [
        ("Experiment", experiment_name),
        ("Scenario", scenario_name),
        ("Export Date", cm.export_date),
        ("Generation Date", datetime.now().isoformat()),
        ("Iterations", cm.total_iterations),
        ("Current Iteration", cm.iteration),
        ("REI", cm.rei),
        ("REI Status", cm.status),
        ("Warning Threshold", cm.thresholds.get("warning", "")),
        ("Critical Threshold", cm.thresholds.get("critical", "")),
        ("Functions Analyzed", len(cm.function_variabilities)),
        ("Resonance Detections", len(cm.resonance_detections)),
        ("Resonance Chains", len(cm.resonance_chains)),
        ("Barrier Functions", len(cm.barrier_functions)),
    ]

    # Header
    ws.cell(row=1, column=1, value="Property")
    ws.cell(row=1, column=2, value="Value")
    _style_header_row(ws, 1, 2)

    for i, (prop, val) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=prop)
        cell = ws.cell(row=i, column=2, value=val)
        cell.border = THIN_BORDER
        ws.cell(row=i, column=1).border = THIN_BORDER

    # Legend
    row = len(data) + 3
    ws.cell(row=row, column=1, value="Cell Color Legend:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    legends = [
        ("White cells", "Raw data from FlowFRAM export", DATA_FILL),
        ("Yellow cells", "Excel FORMULAS (editable, auto-recompute)", FORMULA_FILL),
        ("Green cells", "Final results / totals", RESULT_FILL),
    ]
    for i, (label, desc, fill) in enumerate(legends, start=1):
        r = row + i
        c = ws.cell(row=r, column=1, value=label)
        c.fill = fill
        c.border = THIN_BORDER
        ws.cell(row=r, column=2, value=desc).border = THIN_BORDER


def _build_function_statistics_sheet(
    wb: Workbook,
    cm: ComplexityMetricsExport,
    stats: StatisticsExport | None,
):
    """
    Sheet 2: Function output statistics with CV formula.

    DATA columns: Function ID, Mean (μ), StdDev (σ), Min, Max, N
    FORMULA column: CV = σ/μ (Excel formula: =C{row}/B{row})
    """
    ws = wb.create_sheet("Function Statistics")

    headers = ["Function ID", "Mean (μ)", "StdDev (σ)", "CV (=σ/μ)", "CV %", "Min", "Max", "N", "Phenotype", "Level"]
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Write headers
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # Build data from complexity metrics + optional stats cross-reference
    row = 2
    for fv in cm.function_variabilities:
        ws.cell(row=row, column=1, value=fv.function_id)
        ws.cell(row=row, column=2, value=fv.mean)       # B: Mean
        # Try to get StdDev from stats export, otherwise compute from CV
        std_val = fv.cv * fv.mean if fv.mean != 0 else 0.0
        if stats:
            for var_name, var_stats in stats.variables.items():
                if fv.function_id in var_name:
                    std_val = var_stats.stats.std
                    break
        ws.cell(row=row, column=3, value=std_val)        # C: StdDev

        # FORMULA: CV = StdDev / Mean
        cv_formula = f'=IF(B{row}=0,"N/A",C{row}/B{row})'
        ws.cell(row=row, column=4, value=cv_formula)
        _style_data_cell(ws, row, 4, is_formula=True)

        # FORMULA: CV% = CV × 100
        cv_pct_formula = f'=IF(B{row}=0,"N/A",D{row}*100)'
        ws.cell(row=row, column=5, value=cv_pct_formula)
        _style_data_cell(ws, row, 5, is_formula=True)

        # Min, Max, N from stats
        min_val, max_val, n_val = 0, 0, cm.total_iterations
        if stats:
            for var_name, var_stats in stats.variables.items():
                if fv.function_id in var_name:
                    min_val = var_stats.stats.min
                    max_val = var_stats.stats.max
                    n_val = var_stats.stats.count
                    break

        ws.cell(row=row, column=6, value=min_val)
        ws.cell(row=row, column=7, value=max_val)
        ws.cell(row=row, column=8, value=n_val)
        ws.cell(row=row, column=9, value=fv.precision_phenotype)
        ws.cell(row=row, column=10, value=fv.variability_level)

        # Style data cells
        for col in [1, 2, 3, 6, 7, 8, 9, 10]:
            _style_data_cell(ws, row, col, is_formula=False)

        row += 1

    # Summary row with formulas
    last_data_row = row - 1
    if last_data_row >= 2:
        ws.cell(row=row + 1, column=1, value="AVERAGES")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)

        # FORMULA: Average Mean
        ws.cell(row=row + 1, column=2, value=f"=AVERAGE(B2:B{last_data_row})")
        _style_data_cell(ws, row + 1, 2, is_formula=True)
        _style_result_cell(ws, row + 1, 2)

        # FORMULA: Average StdDev
        ws.cell(row=row + 1, column=3, value=f"=AVERAGE(C2:C{last_data_row})")
        _style_data_cell(ws, row + 1, 3, is_formula=True)
        _style_result_cell(ws, row + 1, 3)

        # FORMULA: Average CV
        ws.cell(row=row + 1, column=4, value=f"=AVERAGE(D2:D{last_data_row})")
        _style_data_cell(ws, row + 1, 4, is_formula=True)
        _style_result_cell(ws, row + 1, 4)

        # FORMULA: Average CV%
        ws.cell(row=row + 1, column=5, value=f"=AVERAGE(E2:E{last_data_row})")
        _style_data_cell(ws, row + 1, 5, is_formula=True)
        _style_result_cell(ws, row + 1, 5)

        # Count
        ws.cell(row=row + 1, column=8, value=f"=SUM(H2:H{last_data_row})")
        _style_data_cell(ws, row + 1, 8, is_formula=True)
        _style_result_cell(ws, row + 1, 8)


def _build_rei_calculation_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 3: REI factor decomposition with formulas.

    DATA columns: Factor Name, Category, Raw Value, Weight (implicit)
    FORMULA columns: Contribution (= value already weighted), REI sum
    """
    ws = wb.create_sheet("REI Calculation")

    headers = [
        "Factor Name", "Category", "Contribution",
        "% of Total (formula)", "Cumulative (formula)"
    ]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    row = 2
    for f in cm.factors:
        ws.cell(row=row, column=1, value=f.name)
        ws.cell(row=row, column=2, value=f.category)
        ws.cell(row=row, column=3, value=f.contribution)  # DATA

        # FORMULA: % of total = contribution / SUM(all contributions)
        last_factor_row = len(cm.factors) + 1
        pct_formula = f"=IF(SUM(C$2:C${last_factor_row})=0,0,C{row}/SUM(C$2:C${last_factor_row}))"
        ws.cell(row=row, column=4, value=pct_formula)
        ws.cell(row=row, column=4).number_format = NUMBER_FMT_PCT
        _style_data_cell(ws, row, 4, is_formula=True)

        # FORMULA: Cumulative sum
        if row == 2:
            cum_formula = f"=C{row}"
        else:
            cum_formula = f"=E{row-1}+C{row}"
        ws.cell(row=row, column=5, value=cum_formula)
        _style_data_cell(ws, row, 5, is_formula=True)

        for col in [1, 2, 3]:
            _style_data_cell(ws, row, col)

        row += 1

    # REI total row
    total_row = row
    ws.cell(row=total_row, column=1, value="REI (Total)")
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)

    # FORMULA: REI = SUM of all contributions
    rei_formula = f"=SUM(C2:C{total_row - 1})"
    ws.cell(row=total_row, column=3, value=rei_formula)
    _style_result_cell(ws, total_row, 3)

    ws.cell(row=total_row, column=4, value="=1")
    ws.cell(row=total_row, column=4).number_format = NUMBER_FMT_PCT
    _style_result_cell(ws, total_row, 4)

    # Static vs Dynamic breakdown
    breakdown_row = total_row + 2
    ws.cell(row=breakdown_row, column=1, value="Static/Dynamic Breakdown")
    _style_subheader_row(ws, breakdown_row, 5)

    ws.cell(row=breakdown_row + 1, column=1, value="Static Contribution")
    static_formula = f'=SUMPRODUCT((B2:B{total_row-1}="static")*C2:C{total_row-1})'
    ws.cell(row=breakdown_row + 1, column=3, value=static_formula)
    _style_data_cell(ws, breakdown_row + 1, 3, is_formula=True)
    _style_result_cell(ws, breakdown_row + 1, 3)

    ws.cell(row=breakdown_row + 2, column=1, value="Dynamic Contribution")
    dynamic_formula = f'=SUMPRODUCT((B2:B{total_row-1}="dynamic")*C2:C{total_row-1})'
    ws.cell(row=breakdown_row + 2, column=3, value=dynamic_formula)
    _style_data_cell(ws, breakdown_row + 2, 3, is_formula=True)
    _style_result_cell(ws, breakdown_row + 2, 3)

    ws.cell(row=breakdown_row + 3, column=1, value="Static %")
    ws.cell(row=breakdown_row + 3, column=3,
            value=f"=IF(C{total_row}=0,0,C{breakdown_row+1}/C{total_row})")
    ws.cell(row=breakdown_row + 3, column=3).number_format = NUMBER_FMT_PCT
    _style_data_cell(ws, breakdown_row + 3, 3, is_formula=True)

    ws.cell(row=breakdown_row + 4, column=1, value="Dynamic %")
    ws.cell(row=breakdown_row + 4, column=3,
            value=f"=IF(C{total_row}=0,0,C{breakdown_row+2}/C{total_row})")
    ws.cell(row=breakdown_row + 4, column=3).number_format = NUMBER_FMT_PCT
    _style_data_cell(ws, breakdown_row + 4, 3, is_formula=True)

    # FlowFRAM cross-check
    check_row = breakdown_row + 6
    ws.cell(row=check_row, column=1, value="FlowFRAM REI (reference)")
    ws.cell(row=check_row, column=3, value=cm.rei)
    _style_data_cell(ws, check_row, 3)

    ws.cell(row=check_row + 1, column=1, value="Spreadsheet REI (formula)")
    ws.cell(row=check_row + 1, column=3, value=f"=C{total_row}")
    _style_data_cell(ws, check_row + 1, 3, is_formula=True)
    _style_result_cell(ws, check_row + 1, 3)

    ws.cell(row=check_row + 2, column=1, value="Difference")
    ws.cell(row=check_row + 2, column=3, value=f"=ABS(C{check_row}-C{check_row+1})")
    _style_data_cell(ws, check_row + 2, 3, is_formula=True)

    ws.cell(row=check_row + 3, column=1, value="Match?")
    ws.cell(row=check_row + 3, column=3,
            value=f'=IF(C{check_row+2}<0.0001,"✓ MATCH","✗ MISMATCH")')
    _style_data_cell(ws, check_row + 3, 3, is_formula=True)


def _build_resonance_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 4: Resonance detections with combined score formula.

    DATA: upstream, downstream, r (correlation), NMI
    FORMULA: CombinedScore = r × (0.5 + 0.5 × NMI)
    """
    ws = wb.create_sheet("Resonance Analysis")

    headers = [
        "Upstream", "Downstream", "Aspect", "Correlation (r)", "NMI",
        "Combined Score (formula)", "Type", "|Score| (formula)"
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18

    row = 2
    for rd in cm.resonance_detections:
        ws.cell(row=row, column=1, value=rd.upstream)
        ws.cell(row=row, column=2, value=rd.downstream)
        ws.cell(row=row, column=3, value=rd.aspect)
        ws.cell(row=row, column=4, value=rd.correlation)         # DATA: r
        ws.cell(row=row, column=5, value=rd.nmi if rd.nmi else 0)  # DATA: NMI

        # FORMULA: Combined Score = r × (0.5 + 0.5 × NMI)
        score_formula = f"=D{row}*(0.5+0.5*E{row})"
        ws.cell(row=row, column=6, value=score_formula)
        _style_data_cell(ws, row, 6, is_formula=True)

        ws.cell(row=row, column=7, value=rd.type)

        # FORMULA: |Score|
        abs_formula = f"=ABS(F{row})"
        ws.cell(row=row, column=8, value=abs_formula)
        _style_data_cell(ws, row, 8, is_formula=True)

        for col in [1, 2, 3, 4, 5, 7]:
            _style_data_cell(ws, row, col)

        row += 1

    # Summary
    last_row = row - 1
    if last_row >= 2:
        sum_row = row + 1
        ws.cell(row=sum_row, column=1, value="SUMMARY")
        ws.cell(row=sum_row, column=1).font = Font(bold=True)

        ws.cell(row=sum_row + 1, column=1, value="Total Detections")
        ws.cell(row=sum_row + 1, column=4, value=f"=COUNTA(D2:D{last_row})")
        _style_data_cell(ws, sum_row + 1, 4, is_formula=True)

        ws.cell(row=sum_row + 2, column=1, value="Avg |Correlation|")
        ws.cell(row=sum_row + 2, column=4, value=f"=AVERAGE(ABS(D2:D{last_row}))")
        _style_data_cell(ws, sum_row + 2, 4, is_formula=True)

        ws.cell(row=sum_row + 3, column=1, value="Avg |Combined Score|")
        ws.cell(row=sum_row + 3, column=6, value=f"=AVERAGE(H2:H{last_row})")
        _style_data_cell(ws, sum_row + 3, 6, is_formula=True)
        _style_result_cell(ws, sum_row + 3, 6)

        ws.cell(row=sum_row + 4, column=1, value="Amplifying Count")
        ws.cell(row=sum_row + 4, column=7,
                value=f'=COUNTIF(G2:G{last_row},"amplifying")')
        _style_data_cell(ws, sum_row + 4, 7, is_formula=True)

        ws.cell(row=sum_row + 5, column=1, value="Dampening Count")
        ws.cell(row=sum_row + 5, column=7,
                value=f'=COUNTIF(G2:G{last_row},"dampening")')
        _style_data_cell(ws, sum_row + 5, 7, is_formula=True)

        # VPI formula
        ws.cell(row=sum_row + 7, column=1, value="VPI (formula)")
        ws.cell(row=sum_row + 7, column=1).font = Font(bold=True)
        vpi_formula = f"=IF(COUNTA(D2:D{last_row})=0,0,SUM(H2:H{last_row})/COUNTA(D2:D{last_row}))"
        ws.cell(row=sum_row + 7, column=6, value=vpi_formula)
        _style_result_cell(ws, sum_row + 7, 6)

        if cm.vpi:
            ws.cell(row=sum_row + 8, column=1, value="FlowFRAM VPI (reference)")
            ws.cell(row=sum_row + 8, column=6, value=cm.vpi.value)
            _style_data_cell(ws, sum_row + 8, 6)

            ws.cell(row=sum_row + 9, column=1, value="Match?")
            ws.cell(row=sum_row + 9, column=6,
                    value=f'=IF(ABS(F{sum_row+7}-F{sum_row+8})<0.001,"✓ MATCH","✗ MISMATCH")')
            _style_data_cell(ws, sum_row + 9, 6, is_formula=True)


def _build_entropy_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 5: Shannon Entropy calculations with formulas.

    Computes aspect entropy and (if data available) coupling entropy.
    FORMULA: H = -Σ p × log2(p)
    """
    ws = wb.create_sheet("Entropy Analysis")

    # Aspect entropy section
    ws.cell(row=1, column=1, value="Aspect Entropy Calculation")
    _style_subheader_row(ws, 1, 6)

    headers = ["Aspect", "Full Name", "Count", "Total (formula)",
               "Probability (formula)", "-p × log₂(p) (formula)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 24

    row = 3
    first_data_row = row
    for av in cm.aspect_variabilities:
        ws.cell(row=row, column=1, value=av.aspect)
        ws.cell(row=row, column=2, value=av.full_name)
        ws.cell(row=row, column=3, value=av.count)  # DATA: count

        # FORMULA: Total = SUM of all counts
        last_aspect_row = first_data_row + len(cm.aspect_variabilities) - 1
        ws.cell(row=row, column=4, value=f"=SUM(C${first_data_row}:C${last_aspect_row})")
        _style_data_cell(ws, row, 4, is_formula=True)

        # FORMULA: Probability = count / total
        ws.cell(row=row, column=5, value=f"=IF(D{row}=0,0,C{row}/D{row})")
        _style_data_cell(ws, row, 5, is_formula=True)

        # FORMULA: -p × log2(p)
        ws.cell(row=row, column=6,
                value=f"=IF(E{row}<=0,0,-E{row}*LOG(E{row},2))")
        _style_data_cell(ws, row, 6, is_formula=True)

        for col in [1, 2, 3]:
            _style_data_cell(ws, row, col)

        row += 1

    # H_aspect = SUM of -p*log2(p)
    h_row = row
    ws.cell(row=h_row, column=1, value="H_aspect")
    ws.cell(row=h_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=h_row, column=6, value=f"=SUM(F{first_data_row}:F{h_row - 1})")
    _style_result_cell(ws, h_row, 6)

    # Max entropy
    ws.cell(row=h_row + 1, column=1, value="H_max (log₂(N))")
    n_aspects = len(cm.aspect_variabilities) if cm.aspect_variabilities else 6
    ws.cell(row=h_row + 1, column=6, value=f"=LOG({n_aspects},2)")
    _style_data_cell(ws, h_row + 1, 6, is_formula=True)

    # Normalized entropy
    ws.cell(row=h_row + 2, column=1, value="H_norm (= H/H_max)")
    ws.cell(row=h_row + 2, column=6,
            value=f"=IF(F{h_row+1}=0,0,F{h_row}/F{h_row+1})")
    _style_data_cell(ws, h_row + 2, 6, is_formula=True)
    _style_result_cell(ws, h_row + 2, 6)


def _build_chains_barriers_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 6: Resonance chains and barrier functions.

    Chain strength = Π|r_i| (product of absolute correlations)
    FORMULA: =PRODUCT(ABS(correlations))
    """
    ws = wb.create_sheet("Chains & Barriers")

    # Chains section
    ws.cell(row=1, column=1, value="Resonance Chains")
    _style_subheader_row(ws, 1, 6)

    headers = ["Chain #", "Path", "Length", "Strength (data)", "Type",
               "Strength Check (formula)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 24

    row = 3
    for i, chain in enumerate(cm.resonance_chains):
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=" → ".join(chain.path))
        ws.cell(row=row, column=3, value=len(chain.path))
        ws.cell(row=row, column=4, value=chain.strength)
        ws.cell(row=row, column=5, value=chain.type)

        # Write individual correlations and build product formula
        # Put correlations in columns G onwards (hidden helper)
        corr_cells = []
        for j, corr in enumerate(chain.correlations):
            corr_col = 7 + j
            ws.cell(row=row, column=corr_col, value=abs(corr))
            corr_cells.append(f"{get_column_letter(corr_col)}{row}")

        if corr_cells:
            product_formula = f"=PRODUCT({','.join(corr_cells)})"
        else:
            product_formula = "=0"
        ws.cell(row=row, column=6, value=product_formula)
        _style_data_cell(ws, row, 6, is_formula=True)

        for col in [1, 2, 3, 4, 5]:
            _style_data_cell(ws, row, col)

        row += 1

    # Barriers section
    barrier_start = row + 2
    ws.cell(row=barrier_start, column=1, value="Barrier Functions")
    _style_subheader_row(ws, barrier_start, 6)

    b_headers = ["Function ID", "Incoming r", "Outgoing r",
                 "Damping Ratio (formula)", "CV", "Effectiveness"]
    for col, h in enumerate(b_headers, start=1):
        ws.cell(row=barrier_start + 1, column=col, value=h)
    _style_header_row(ws, barrier_start + 1, len(b_headers))

    row = barrier_start + 2
    for bf in cm.barrier_functions:
        ws.cell(row=row, column=1, value=bf.function_id)
        ws.cell(row=row, column=2, value=bf.incoming_correlation)
        ws.cell(row=row, column=3, value=bf.outgoing_correlation)

        # FORMULA: Damping ratio = 1 - (outgoing / incoming)
        damp_formula = f"=IF(B{row}=0,0,1-(ABS(C{row})/ABS(B{row})))"
        ws.cell(row=row, column=4, value=damp_formula)
        _style_data_cell(ws, row, 4, is_formula=True)

        ws.cell(row=row, column=5, value=bf.cv)
        ws.cell(row=row, column=6, value=bf.effectiveness)

        for col in [1, 2, 3, 5, 6]:
            _style_data_cell(ws, row, col)

        row += 1


def _build_itd_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 7: Dialogical Tension Index (ITD) with formulas.

    ITD = -Σ p_s × log₂(p_s) / log₂(|S|)
    where S = {success, indeterminate, failure}
    """
    ws = wb.create_sheet("ITD Analysis")

    # First: function outputs for threshold computation
    ws.cell(row=1, column=1, value="ITD (Dialogical Tension Index) — Morin Dialogic Principle")
    _style_subheader_row(ws, 1, 8)

    # Threshold computation
    ws.cell(row=3, column=1, value="Threshold Computation (Statistical Mode: mean ± σ)")
    ws.cell(row=3, column=1).font = Font(bold=True)

    ws.cell(row=4, column=1, value="Grand Mean of outputs")
    ws.cell(row=4, column=2, value="Grand σ of outputs")
    ws.cell(row=4, column=3, value="Success Threshold (≥)")
    ws.cell(row=4, column=4, value="Failure Threshold (≤)")
    ws.cell(row=4, column=5, value="log₂(3)")
    _style_header_row(ws, 4, 5)

    # Collect means for threshold computation
    means = [fv.mean for fv in cm.function_variabilities if fv.mean != 0]
    grand_mean = sum(means) / len(means) if means else 0
    grand_std = (sum((m - grand_mean)**2 for m in means) / max(len(means) - 1, 1)) ** 0.5 if len(means) > 1 else 0

    ws.cell(row=5, column=1, value=grand_mean)
    ws.cell(row=5, column=2, value=grand_std)
    # FORMULA: Success = mean + σ
    ws.cell(row=5, column=3, value="=A5+B5")
    _style_data_cell(ws, 5, 3, is_formula=True)
    # FORMULA: Failure = mean - σ
    ws.cell(row=5, column=4, value="=A5-B5")
    _style_data_cell(ws, 5, 4, is_formula=True)
    # FORMULA: log₂(3)
    ws.cell(row=5, column=5, value="=LOG(3,2)")
    _style_data_cell(ws, 5, 5, is_formula=True)

    # Per-function ITD
    ws.cell(row=7, column=1, value="Per-Function ITD")
    _style_subheader_row(ws, 7, 8)

    headers = ["Function", "Mean (μ)", "CV", "σ (=μ×CV)",
               "P(Success)", "P(Failure)", "P(Indet.)",
               "ITD_norm (formula)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=8, column=col, value=h)
    _style_header_row(ws, 8, len(headers))

    ws.column_dimensions["A"].width = 24
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16

    row = 9
    first = row
    for fv in cm.function_variabilities:
        ws.cell(row=row, column=1, value=fv.function_id)
        ws.cell(row=row, column=2, value=fv.mean)   # DATA
        ws.cell(row=row, column=3, value=fv.cv)      # DATA

        # FORMULA: σ = μ × CV
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        _style_data_cell(ws, row, 4, is_formula=True)

        # Note: Exact P(Success/Failure) for normal distribution requires
        # NORM.DIST which is available in Excel. Using approximation approach.
        # FORMULA: P(Success) = 1 - NORM.DIST(success_threshold, μ, σ, TRUE)
        ws.cell(row=row, column=5,
                value=f'=IF(D{row}<=0,IF(B{row}>=$C$5,1,0),1-NORM.DIST($C$5,B{row},D{row},TRUE))')
        _style_data_cell(ws, row, 5, is_formula=True)

        # FORMULA: P(Failure) = NORM.DIST(failure_threshold, μ, σ, TRUE)
        ws.cell(row=row, column=6,
                value=f'=IF(D{row}<=0,IF(B{row}<=$D$5,1,0),NORM.DIST($D$5,B{row},D{row},TRUE))')
        _style_data_cell(ws, row, 6, is_formula=True)

        # FORMULA: P(Indeterminate) = 1 - P(S) - P(F)
        ws.cell(row=row, column=7, value=f"=MAX(0,1-E{row}-F{row})")
        _style_data_cell(ws, row, 7, is_formula=True)

        # FORMULA: ITD_norm = (-Σ p*log2(p)) / log2(3)
        # = (-IF(E>0,E*LOG(E,2),0) - IF(F>0,F*LOG(F,2),0) - IF(G>0,G*LOG(G,2),0)) / log2(3)
        itd_formula = (
            f"=(IF(E{row}>0,-E{row}*LOG(E{row},2),0)"
            f"+IF(F{row}>0,-F{row}*LOG(F{row},2),0)"
            f"+IF(G{row}>0,-G{row}*LOG(G{row},2),0))/$E$5"
        )
        ws.cell(row=row, column=8, value=itd_formula)
        _style_data_cell(ws, row, 8, is_formula=True)

        for col in [1, 2, 3]:
            _style_data_cell(ws, row, col)

        row += 1

    # System ITD
    last = row - 1
    ws.cell(row=row + 1, column=1, value="System ITD (average)")
    ws.cell(row=row + 1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row + 1, column=8, value=f"=AVERAGE(H{first}:H{last})")
    _style_result_cell(ws, row + 1, 8)


def _build_cri_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 8: CRI — Criticality Ranking Index.

    Ranks functions by normalised CV contribution.
    DATA: Function ID, CV, Mean
    FORMULA: NormScore = CV_i / SUM(CV), CumulativeRank
    """
    ws = wb.create_sheet("CRI Ranking")

    headers = [
        "Rank", "Function ID", "CV", "Mean (μ)",
        "Norm Score (formula)", "Cum. Score (formula)", "Percentile (formula)",
    ]
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    for c in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    # Sort by CV descending
    sorted_fvs = sorted(cm.function_variabilities, key=lambda fv: fv.cv, reverse=True)

    row = 2
    first, last = row, row + len(sorted_fvs) - 1
    for i, fv in enumerate(sorted_fvs):
        ws.cell(row=row, column=1, value=i + 1)                  # Rank
        ws.cell(row=row, column=2, value=fv.function_id)         # Function
        ws.cell(row=row, column=3, value=fv.cv)                  # CV (DATA)
        ws.cell(row=row, column=4, value=fv.mean)                # Mean (DATA)

        # FORMULA: NormScore = CV_i / SUM(all CVs)
        ws.cell(row=row, column=5,
                value=f"=IF(SUM(C${first}:C${last})=0,0,C{row}/SUM(C${first}:C${last}))")
        _style_data_cell(ws, row, 5, is_formula=True)

        # FORMULA: Cumulative score
        if row == first:
            ws.cell(row=row, column=6, value=f"=E{row}")
        else:
            ws.cell(row=row, column=6, value=f"=F{row-1}+E{row}")
        _style_data_cell(ws, row, 6, is_formula=True)

        # FORMULA: Percentile (rank / N)
        ws.cell(row=row, column=7,
                value=f"=A{row}/{last - first + 1}")
        ws.cell(row=row, column=7).number_format = NUMBER_FMT_PCT
        _style_data_cell(ws, row, 7, is_formula=True)

        for col in [1, 2, 3, 4]:
            _style_data_cell(ws, row, col)

        row += 1

    # Summary
    if sorted_fvs:
        ws.cell(row=row + 1, column=2, value="Top Contributor")
        ws.cell(row=row + 1, column=2).font = Font(bold=True)
        ws.cell(row=row + 1, column=3, value=f"=B{first}")
        _style_result_cell(ws, row + 1, 3)

        ws.cell(row=row + 2, column=2, value="Top CV")
        ws.cell(row=row + 2, column=3, value=f"=C{first}")
        _style_result_cell(ws, row + 2, 3)

        ws.cell(row=row + 3, column=2, value="Bottom CV")
        ws.cell(row=row + 3, column=3, value=f"=C{last}")
        _style_result_cell(ws, row + 3, 3)

        # Gini-like concentration: top 20% share
        top20_count = max(1, len(sorted_fvs) // 5)
        top20_last = first + top20_count - 1
        ws.cell(row=row + 4, column=2, value="Top 20% CV Share")
        ws.cell(row=row + 4, column=3,
                value=f"=IF(SUM(C${first}:C${last})=0,0,SUM(C${first}:C${top20_last})/SUM(C${first}:C${last}))")
        ws.cell(row=row + 4, column=3).number_format = NUMBER_FMT_PCT
        _style_data_cell(ws, row + 4, 3, is_formula=True)
        _style_result_cell(ws, row + 4, 3)


def _build_entropy_sync_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 9: Entropy Synchronization 2.0 (ES 2.0).

    Displays Bellman-Picard convergence data and amplification factor.
    DATA: Direct from cm.entropy_synchronization dict.
    """
    ws = wb.create_sheet("ES 2.0")

    ws.cell(row=1, column=1, value="Entropy Synchronization 2.0 — Bellman-Picard Convergence")
    _style_subheader_row(ws, 1, 4)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20

    es = cm.entropy_synchronization or {}

    # Properties section
    props = [
        ("Convergence Method", "Bellman-Picard" if es.get("wasRecursive") else "Direct"),
        ("Was Recursive", str(es.get("wasRecursive", False))),
        ("Convergence Iterations", es.get("convergenceIterations", "N/A")),
        ("Convergence Residual", es.get("convergenceResidual", "N/A")),
        ("Amplification Factor", es.get("amplificationFactor", "N/A")),
        ("Max Iterations (config)", es.get("maxIterations", "N/A")),
        ("Epsilon (config)", es.get("epsilon", "N/A")),
        ("Damping (config)", es.get("damping", "N/A")),
        ("Participating Functions", es.get("participatingFunctions", "N/A")),
    ]

    ws.cell(row=3, column=1, value="Property")
    ws.cell(row=3, column=2, value="Value")
    _style_header_row(ws, 3, 2)

    for i, (prop, val) in enumerate(props, start=4):
        ws.cell(row=i, column=1, value=prop)
        ws.cell(row=i, column=2, value=val)
        for c in [1, 2]:
            _style_data_cell(ws, i, c)

    # Amplification analysis
    amp_row = 4 + len(props) + 1
    ws.cell(row=amp_row, column=1, value="Amplification Analysis")
    _style_subheader_row(ws, amp_row, 4)

    amp = es.get("amplificationFactor", 1.0)
    if isinstance(amp, (int, float)):
        ws.cell(row=amp_row + 1, column=1, value="Amplification Factor")
        ws.cell(row=amp_row + 1, column=2, value=amp)
        _style_data_cell(ws, amp_row + 1, 2)

        ws.cell(row=amp_row + 2, column=1, value="Interpretation")
        if amp > 1.1:
            interp = "Significant amplification — network effects increase entropy"
        elif amp > 1.0:
            interp = "Minor amplification — small network effect"
        elif amp == 1.0:
            interp = "No amplification — entropy is self-consistent"
        else:
            interp = "Dampening — network effects reduce entropy"
        ws.cell(row=amp_row + 2, column=2, value=interp)
        _style_data_cell(ws, amp_row + 2, 2)

    # Description
    desc = es.get("description", "")
    if desc:
        desc_row = amp_row + 4
        ws.cell(row=desc_row, column=1, value="FlowFRAM Description")
        ws.cell(row=desc_row, column=1).font = Font(bold=True)
        ws.cell(row=desc_row + 1, column=1, value=desc)


def _build_rei_nonlinear_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 10: REI Non-Linear Composition.

    Shows FOV, FRC, ES linear values and their pairwise interaction terms.
    FORMULA: NL-REI = linear + interactions
    """
    ws = wb.create_sheet("REI Non-Linear")

    ws.cell(row=1, column=1, value="REI Non-Linear Composition — Factor Interaction Analysis")
    _style_subheader_row(ws, 1, 4)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    rc = cm.rei_composition or {}
    factor_vals = rc.get("factorValues", {})
    interactions = rc.get("interactionTerms", {})

    # Factor values section
    ws.cell(row=3, column=1, value="Factor")
    ws.cell(row=3, column=2, value="Value")
    _style_header_row(ws, 3, 2)

    factor_data = [
        ("FOV (Function Output Variability)", factor_vals.get("FOV", 0)),
        ("FRC (Functional Resonance Coupling)", factor_vals.get("FRC", 0)),
        ("ES (Entropy Synchronization)", factor_vals.get("ES", 0)),
    ]
    for i, (name, val) in enumerate(factor_data, start=4):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=val)
        _style_data_cell(ws, i, 1)
        _style_data_cell(ws, i, 2)

    # Linear REI
    lin_row = 4 + len(factor_data) + 1
    ws.cell(row=lin_row, column=1, value="Linear REI")
    ws.cell(row=lin_row, column=1).font = Font(bold=True)
    ws.cell(row=lin_row, column=2, value=rc.get("linearREI", cm.rei))
    _style_data_cell(ws, lin_row, 2)

    # FORMULA: Sum factors to verify
    ws.cell(row=lin_row, column=3, value="=B4+B5+B6")
    ws.cell(row=lin_row, column=3).font = MONO_FONT
    _style_data_cell(ws, lin_row, 3, is_formula=True)

    # Interaction terms section
    int_row = lin_row + 2
    ws.cell(row=int_row, column=1, value="Interaction Terms")
    _style_subheader_row(ws, int_row, 4)

    ws.cell(row=int_row + 1, column=1, value="Term")
    ws.cell(row=int_row + 1, column=2, value="Value")
    ws.cell(row=int_row + 1, column=3, value="Formula Check")
    _style_header_row(ws, int_row + 1, 3)

    int_data = [
        ("FOV × FRC", interactions.get("FOV_x_FRC", 0), "=B4*B5"),
        ("FOV × ES", interactions.get("FOV_x_ES", 0), "=B4*B6"),
        ("FRC × ES", interactions.get("FRC_x_ES", 0), "=B5*B6"),
    ]
    for i, (name, val, formula) in enumerate(int_data, start=int_row + 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=val)
        _style_data_cell(ws, i, 1)
        _style_data_cell(ws, i, 2)
        ws.cell(row=i, column=3, value=formula)
        _style_data_cell(ws, i, 3, is_formula=True)

    total_int_row = int_row + 2 + len(int_data)
    ws.cell(row=total_int_row, column=1, value="Total Interactions")
    ws.cell(row=total_int_row, column=1).font = Font(bold=True)
    ws.cell(row=total_int_row, column=2, value=interactions.get("total", 0))
    _style_data_cell(ws, total_int_row, 2)

    first_int = int_row + 2
    last_int = first_int + len(int_data) - 1
    ws.cell(row=total_int_row, column=3,
            value=f"=SUM(B{first_int}:B{last_int})")
    _style_data_cell(ws, total_int_row, 3, is_formula=True)

    # Non-linear REI
    nl_row = total_int_row + 2
    ws.cell(row=nl_row, column=1, value="Non-Linear REI")
    ws.cell(row=nl_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=nl_row, column=2, value=rc.get("nonLinearREI", cm.rei))
    _style_result_cell(ws, nl_row, 2)

    # FORMULA: NL-REI = Linear + Interactions
    ws.cell(row=nl_row, column=3,
            value=f"=B{lin_row}+B{total_int_row}")
    _style_data_cell(ws, nl_row, 3, is_formula=True)
    _style_result_cell(ws, nl_row, 3)

    # Interaction percentage
    ws.cell(row=nl_row + 1, column=1, value="Interaction % of Total")
    ws.cell(row=nl_row + 1, column=2, value=rc.get("interactionPercent", 0))
    ws.cell(row=nl_row + 1, column=2).number_format = NUMBER_FMT_4
    _style_data_cell(ws, nl_row + 1, 2)

    # FORMULA: Interaction % = total_int / NL-REI
    ws.cell(row=nl_row + 1, column=3,
            value=f"=IF(B{nl_row}=0,0,B{total_int_row}/B{nl_row}*100)")
    ws.cell(row=nl_row + 1, column=3).number_format = NUMBER_FMT_4
    _style_data_cell(ws, nl_row + 1, 3, is_formula=True)

    # Cross-check with FlowFRAM
    chk_row = nl_row + 3
    ws.cell(row=chk_row, column=1, value="FlowFRAM NL-REI (reference)")
    ws.cell(row=chk_row, column=2, value=rc.get("nonLinearREI", cm.rei))
    _style_data_cell(ws, chk_row, 2)
    ws.cell(row=chk_row + 1, column=1, value="Spreadsheet NL-REI")
    ws.cell(row=chk_row + 1, column=2, value=f"=C{nl_row}")
    _style_data_cell(ws, chk_row + 1, 2, is_formula=True)
    ws.cell(row=chk_row + 2, column=1, value="Match?")
    ws.cell(row=chk_row + 2, column=2,
            value=f'=IF(ABS(B{chk_row}-B{chk_row+1})<0.001,"✓ MATCH","✗ MISMATCH")')
    _style_data_cell(ws, chk_row + 2, 2, is_formula=True)


def _build_entropy_rate_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 11: Entropy Rate — per-function convergence trends.

    DATA: functionId, currentCV, previousCV
    FORMULA: rate = currentCV - previousCV; trend classification
    """
    ws = wb.create_sheet("Entropy Rate")

    ws.cell(row=1, column=1, value="Entropy Rate — Per-Function Convergence Trends")
    _style_subheader_row(ws, 1, 7)

    ws.column_dimensions["A"].width = 28
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 18

    er = cm.entropy_rate or {}
    per_fn = er.get("perFunction", [])

    # System summary
    ws.cell(row=3, column=1, value="System Entropy Rate")
    ws.cell(row=3, column=2, value=er.get("systemRate", 0))
    ws.cell(row=3, column=1).font = Font(bold=True)
    _style_data_cell(ws, 3, 2)

    ws.cell(row=4, column=1, value="System Trend")
    ws.cell(row=4, column=2, value=er.get("systemTrend", "N/A"))
    _style_data_cell(ws, 4, 2)

    trends = er.get("trends", {})
    ws.cell(row=5, column=1, value="Trends Breakdown")
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=5, column=2, value=f"↑ {trends.get('increasing', 0)}")
    ws.cell(row=5, column=3, value=f"↓ {trends.get('decreasing', 0)}")
    ws.cell(row=5, column=4, value=f"→ {trends.get('stable', 0)}")
    ws.cell(row=5, column=5, value=f"NEW {trends.get('new', 0)}")

    # Per-function table
    ws.cell(row=7, column=1, value="Per-Function Details")
    _style_subheader_row(ws, 7, 7)

    headers = [
        "Function ID", "Current CV", "Previous CV",
        "Rate (formula)", "Trend", "|Rate| (formula)", "Converging? (formula)"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=8, column=col, value=h)
    _style_header_row(ws, 8, len(headers))

    row = 9
    first = row
    for fn in per_fn:
        ws.cell(row=row, column=1, value=fn.get("functionId", ""))
        ws.cell(row=row, column=2, value=fn.get("currentCV", 0))     # B: DATA
        ws.cell(row=row, column=3, value=fn.get("previousCV", 0))    # C: DATA

        # FORMULA: Rate = current - previous
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        _style_data_cell(ws, row, 4, is_formula=True)

        ws.cell(row=row, column=5, value=fn.get("trend", ""))

        # FORMULA: |Rate|
        ws.cell(row=row, column=6, value=f"=ABS(D{row})")
        _style_data_cell(ws, row, 6, is_formula=True)

        # FORMULA: Converging if rate < 0 (CV decreasing)
        ws.cell(row=row, column=7,
                value=f'=IF(D{row}<-0.001,"✓ Converging",IF(D{row}>0.001,"⚠ Diverging","→ Stable"))')
        _style_data_cell(ws, row, 7, is_formula=True)

        for col in [1, 2, 3, 5]:
            _style_data_cell(ws, row, col)

        row += 1

    # Summary formulas
    last = row - 1
    if last >= first:
        ws.cell(row=row + 1, column=1, value="Avg |Rate|")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        ws.cell(row=row + 1, column=6, value=f"=AVERAGE(F{first}:F{last})")
        _style_result_cell(ws, row + 1, 6)

        # FORMULA: Count converging
        ws.cell(row=row + 2, column=1, value="Converging Count")
        ws.cell(row=row + 2, column=7,
                value=f'=COUNTIF(G{first}:G{last},"✓*")')
        _style_data_cell(ws, row + 2, 7, is_formula=True)

        ws.cell(row=row + 3, column=1, value="Diverging Count")
        ws.cell(row=row + 3, column=7,
                value=f'=COUNTIF(G{first}:G{last},"⚠*")')
        _style_data_cell(ws, row + 3, 7, is_formula=True)

        # Cross-check system rate
        ws.cell(row=row + 5, column=1, value="FlowFRAM System Rate (ref)")
        ws.cell(row=row + 5, column=4, value=er.get("systemRate", 0))
        _style_data_cell(ws, row + 5, 4)

        ws.cell(row=row + 6, column=1, value="Spreadsheet Avg Rate")
        ws.cell(row=row + 6, column=4, value=f"=AVERAGE(D{first}:D{last})")
        _style_data_cell(ws, row + 6, 4, is_formula=True)
        _style_result_cell(ws, row + 6, 4)


def _build_transfer_entropy_sheet(wb: Workbook, cm: ComplexityMetricsExport):
    """
    Sheet 12: Transfer Entropy — per-coupling causal analysis.

    DATA: upstream, downstream, forward TE, reverse TE
    FORMULA: net TE = forward - reverse; causal direction
    """
    ws = wb.create_sheet("Transfer Entropy")

    ws.cell(row=1, column=1, value="Transfer Entropy — Causal Information Flow Analysis")
    _style_subheader_row(ws, 1, 9)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    for c in range(4, 10):
        ws.column_dimensions[get_column_letter(c)].width = 18

    te = cm.transfer_entropy or {}
    couplings = te.get("couplings", [])
    te_summary = te.get("summary", {})

    headers = [
        "Upstream", "Downstream", "Aspect", "Forward TE",
        "Reverse TE", "Net TE (formula)", "Direction (formula)",
        "Confidence", "|Net TE| (formula)"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    row = 3
    first = row
    for c in couplings:
        ws.cell(row=row, column=1, value=c.get("upstream", ""))
        ws.cell(row=row, column=2, value=c.get("downstream", ""))
        ws.cell(row=row, column=3, value=c.get("aspect", ""))
        ws.cell(row=row, column=4, value=c.get("forward", 0))          # DATA
        ws.cell(row=row, column=5, value=c.get("reverse", 0))          # DATA

        # FORMULA: Net TE = Forward - Reverse
        ws.cell(row=row, column=6, value=f"=D{row}-E{row}")
        _style_data_cell(ws, row, 6, is_formula=True)

        # FORMULA: Direction
        ws.cell(row=row, column=7,
                value=f'=IF(ABS(F{row})<0.01,"symmetric",IF(F{row}>0,"confirmed","reversed"))')
        _style_data_cell(ws, row, 7, is_formula=True)

        ws.cell(row=row, column=8, value=c.get("confidence", ""))

        # FORMULA: |Net TE|
        ws.cell(row=row, column=9, value=f"=ABS(F{row})")
        _style_data_cell(ws, row, 9, is_formula=True)

        for col in [1, 2, 3, 4, 5, 8]:
            _style_data_cell(ws, row, col)

        row += 1

    # Summary section
    last = row - 1
    sum_row = row + 1
    ws.cell(row=sum_row, column=1, value="SUMMARY")
    ws.cell(row=sum_row, column=1).font = Font(bold=True, size=12)

    if last >= first:
        ws.cell(row=sum_row + 1, column=1, value="Total Couplings")
        ws.cell(row=sum_row + 1, column=4, value=f"=COUNTA(D{first}:D{last})")
        _style_data_cell(ws, sum_row + 1, 4, is_formula=True)

        ws.cell(row=sum_row + 2, column=1, value="Avg Forward TE")
        ws.cell(row=sum_row + 2, column=4, value=f"=AVERAGE(D{first}:D{last})")
        _style_data_cell(ws, sum_row + 2, 4, is_formula=True)

        ws.cell(row=sum_row + 3, column=1, value="Avg Reverse TE")
        ws.cell(row=sum_row + 3, column=5, value=f"=AVERAGE(E{first}:E{last})")
        _style_data_cell(ws, sum_row + 3, 5, is_formula=True)

        ws.cell(row=sum_row + 4, column=1, value="Avg Net TE")
        ws.cell(row=sum_row + 4, column=6, value=f"=AVERAGE(F{first}:F{last})")
        _style_data_cell(ws, sum_row + 4, 6, is_formula=True)
        _style_result_cell(ws, sum_row + 4, 6)

        # Causal direction counts
        ws.cell(row=sum_row + 5, column=1, value="Causal Confirmed")
        ws.cell(row=sum_row + 5, column=7,
                value=f'=COUNTIF(G{first}:G{last},"confirmed")')
        _style_data_cell(ws, sum_row + 5, 7, is_formula=True)

        ws.cell(row=sum_row + 6, column=1, value="Causal Reversed")
        ws.cell(row=sum_row + 6, column=7,
                value=f'=COUNTIF(G{first}:G{last},"reversed")')
        _style_data_cell(ws, sum_row + 6, 7, is_formula=True)

        ws.cell(row=sum_row + 7, column=1, value="Symmetric")
        ws.cell(row=sum_row + 7, column=7,
                value=f'=COUNTIF(G{first}:G{last},"symmetric")')
        _style_data_cell(ws, sum_row + 7, 7, is_formula=True)

    # Cross-check with FlowFRAM summary
    chk_row = sum_row + 9
    ws.cell(row=chk_row, column=1, value="FlowFRAM TE Summary (reference)")
    ws.cell(row=chk_row, column=1).font = Font(bold=True)

    ref_data = [
        ("Total Couplings", te_summary.get("totalCouplings", 0)),
        ("Causal Confirmed", te_summary.get("causalConfirmed", 0)),
        ("Causal Reversed", te_summary.get("causalReversed", 0)),
        ("Symmetric", te_summary.get("symmetric", 0)),
        ("Avg Forward TE", te_summary.get("avgForwardTE", 0)),
        ("Avg Reverse TE", te_summary.get("avgReverseTE", 0)),
        ("Avg Net TE", te_summary.get("avgNetTE", 0)),
        ("Confidence", te_summary.get("confidence", "N/A")),
    ]
    for i, (label, val) in enumerate(ref_data, start=chk_row + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
        _style_data_cell(ws, i, 1)
        _style_data_cell(ws, i, 2)


def _build_variable_stats_sheet(
    wb: Workbook,
    stats: StatisticsExport,
):
    """
    Sheet 13 (optional): Full variable statistics from statistics export.

    DATA: All descriptive statistics per variable
    FORMULA: CV = std/mean, Range = max-min
    """
    ws = wb.create_sheet("Variable Statistics")

    headers = [
        "Variable", "N", "Mean", "StdDev", "CV (formula)", "CV% (formula)",
        "Min", "Max", "Range (formula)", "P5", "P25", "Median (P50)", "P75", "P95"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    ws.column_dimensions["A"].width = 40
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    row = 2
    for var_name, vs in stats.variables.items():
        ws.cell(row=row, column=1, value=vs.variable_name or var_name)
        ws.cell(row=row, column=2, value=vs.stats.count)
        ws.cell(row=row, column=3, value=vs.stats.mean)     # C
        ws.cell(row=row, column=4, value=vs.stats.std)       # D

        # FORMULA: CV = std / mean
        ws.cell(row=row, column=5, value=f'=IF(C{row}=0,"N/A",D{row}/C{row})')
        _style_data_cell(ws, row, 5, is_formula=True)

        # FORMULA: CV% = CV × 100
        ws.cell(row=row, column=6, value=f'=IF(C{row}=0,"N/A",E{row}*100)')
        _style_data_cell(ws, row, 6, is_formula=True)

        ws.cell(row=row, column=7, value=vs.stats.min)       # G
        ws.cell(row=row, column=8, value=vs.stats.max)       # H

        # FORMULA: Range = max - min
        ws.cell(row=row, column=9, value=f"=H{row}-G{row}")
        _style_data_cell(ws, row, 9, is_formula=True)

        ws.cell(row=row, column=10, value=vs.stats.p5)
        ws.cell(row=row, column=11, value=vs.stats.p25)
        ws.cell(row=row, column=12, value=vs.stats.p50)
        ws.cell(row=row, column=13, value=vs.stats.p75)
        ws.cell(row=row, column=14, value=vs.stats.p95)

        for col in [1, 2, 3, 4, 7, 8, 10, 11, 12, 13, 14]:
            _style_data_cell(ws, row, col)

        row += 1


# ============================================================================
# MAIN GENERATOR
# ============================================================================

def generate_scenario_xlsx(
    experiment_name: str,
    scenario_name: str,
    cm: ComplexityMetricsExport,
    stats: StatisticsExport | None = None,
    output_path: str = "output.xlsx",
) -> str:
    """
    Generate a complete XLSX workbook with live formulas for one scenario.

    Args:
        experiment_name: e.g. "Patriarca et al. (2024)"
        scenario_name: e.g. "S0 — Nominal Baseline"
        cm: Parsed complexity-metrics export
        stats: Parsed statistics export (optional, enriches data)
        output_path: Path to write the XLSX file

    Returns:
        Path to the generated file.
    """
    wb = Workbook()

    # Build all sheets
    _build_metadata_sheet(wb, experiment_name, scenario_name, cm)
    _build_function_statistics_sheet(wb, cm, stats)
    _build_rei_calculation_sheet(wb, cm)
    _build_resonance_sheet(wb, cm)
    _build_entropy_sheet(wb, cm)
    _build_chains_barriers_sheet(wb, cm)
    _build_itd_sheet(wb, cm)
    _build_cri_sheet(wb, cm)

    # Epistemological extension sheets (only when data available)
    if cm.entropy_synchronization:
        _build_entropy_sync_sheet(wb, cm)
    if cm.rei_composition:
        _build_rei_nonlinear_sheet(wb, cm)
    if cm.entropy_rate:
        _build_entropy_rate_sheet(wb, cm)
    if cm.transfer_entropy:
        _build_transfer_entropy_sheet(wb, cm)

    # Variable statistics sheet (only if stats export available)
    if stats and stats.variables:
        _build_variable_stats_sheet(wb, stats)

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def generate_comparison_xlsx(
    experiment_name: str,
    scenarios: list[tuple[str, ComplexityMetricsExport]],
    output_path: str = "comparison.xlsx",
) -> str:
    """
    Generate a cross-scenario comparison spreadsheet.

    Creates a summary sheet where each row is a scenario and columns
    are key metrics, with formulas for deltas between scenarios.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario Comparison"

    headers = [
        "Scenario", "REI", "Status", "# Functions", "# Resonances",
        "# Chains", "# Barriers", "Avg CV%", "Max CV%",
        "VPI", "ΔREI vs S0 (formula)"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    ws.column_dimensions["A"].width = 40
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    for row_idx, (scenario_name, cm) in enumerate(scenarios, start=2):
        ws.cell(row=row_idx, column=1, value=scenario_name)
        ws.cell(row=row_idx, column=2, value=cm.rei)
        ws.cell(row=row_idx, column=3, value=cm.status)
        ws.cell(row=row_idx, column=4, value=len(cm.function_variabilities))
        ws.cell(row=row_idx, column=5, value=len(cm.resonance_detections))
        ws.cell(row=row_idx, column=6, value=len(cm.resonance_chains))
        ws.cell(row=row_idx, column=7, value=len(cm.barrier_functions))

        # FORMULA: Avg CV%
        cvs = [fv.cv_percentage for fv in cm.function_variabilities]
        avg_cv = sum(cvs) / len(cvs) if cvs else 0
        ws.cell(row=row_idx, column=8, value=avg_cv)

        # Max CV%
        max_cv = max(cvs) if cvs else 0
        ws.cell(row=row_idx, column=9, value=max_cv)

        # VPI
        ws.cell(row=row_idx, column=10, value=cm.vpi.value if cm.vpi else 0)

        # FORMULA: ΔREI vs first scenario (S0/Baseline)
        delta_formula = f"=B{row_idx}-B$2"
        ws.cell(row=row_idx, column=11, value=delta_formula)
        _style_data_cell(ws, row_idx, 11, is_formula=True)

        for col in range(1, 11):
            _style_data_cell(ws, row_idx, col)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
